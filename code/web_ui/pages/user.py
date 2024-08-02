import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill
from io import BytesIO
from utils.menu import menu_with_redirect
from utils.langdetect import detect_language_of
import time
from utils.utils import (
    list_dictionary_ids,
    list_supported_language_codes,
    list_translate_models,
    translate_content,
)

st.set_page_config(
    page_title="File Translate",
    page_icon="🚎",
)

# 全局常量
TARGET_LANG = 'zh-cn'

# 获取可用的字典、模型和支持的语言代码列表
model_list = list_translate_models()
dictionaries = list_dictionary_ids() or ['default_dictionary']
supported_lang_codes_dict = list_supported_language_codes()

def init_streamlit():
    """初始化 Streamlit 界面"""
    menu_with_redirect()
    st.title("Excel 文件语言处理器")
    st.markdown(f"您当前以 {st.session_state.role} 角色登录。")

def is_not_number(text):
    """检查文本是否不是数字"""
    try:
        float(text)
        return False
    except ValueError:
        return True

def process_excel(file, target_lang):
    """处理Excel文件，标记非目标语言的单元格，并收集统计信息"""
    start_time = time.time()
    workbook = openpyxl.load_workbook(file)
    progress_bar = st.progress(0)
    total_sheets = len(workbook.sheetnames)
    
    st.write(f"Excel文件共有 {total_sheets} 个工作表")
    
    total_rows = 0
    total_cells = 0
    non_target_cells = 0

    for sheet_index, sheet_name in enumerate(workbook.sheetnames):
        sheet = workbook[sheet_name]
        sheet_rows = sheet.max_row
        sheet_cells = sheet.max_column * sheet_rows
        total_rows += sheet_rows
        total_cells += sheet_cells
        
        st.write(f"处理工作表: {sheet_name} (行数: {sheet_rows}, 单元格数: {sheet_cells})")
        
        for row_index, row in enumerate(sheet.iter_rows(), 1):
            for cell in row:
                if cell.data_type == 's' and is_not_number(cell.value):
                    lang = detect_language_of(cell.value)
                    if lang != target_lang:
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        non_target_cells += 1
            
            # 更新进度条
            progress = (sheet_index * sheet_rows + row_index) / (total_sheets * sheet_rows)
            progress_bar.progress(progress)

    end_time = time.time()
    processing_time = end_time - start_time

    buffer = BytesIO()
    workbook.save(buffer)
    
    return buffer.getvalue(), {
        "total_sheets": total_sheets,
        "total_rows": total_rows,
        "total_cells": total_cells,
        "non_target_cells": non_target_cells,
        "processing_time": processing_time
    }

def upload_file():
    """处理文件上传"""
    return st.file_uploader("选择一个Excel文件", type=["xlsx", "xls"])

def main():
    init_streamlit()
    
    uploaded_file = upload_file()
    
    if uploaded_file is not None:
        st.write("文件名:", uploaded_file.name)
        
        dictionary_name = st.selectbox(
            "专词映射表", 
            dictionaries, 
            index=dictionaries.index('anthropic.claude-3-haiku-20240307-v1:0') if 'anthropic.claude-3-haiku-20240307-v1:0' in dictionaries else 0
        )
        model_id = st.selectbox("翻译模型", model_list)

        supported_lang_codes = supported_lang_codes_dict.keys() 

        target_lang_label = st.selectbox(
            "目标语言", 
            supported_lang_codes, 
            index=supported_lang_codes.index('zh-cn') if 'zh-cn' in supported_lang_codes else 0
        )

        target_lang = supported_lang_codes_dict.get(target_lang_label)
        
        if st.button("处理文件"):
            with st.spinner('处理中...'):
                processed_data, stats = process_excel(uploaded_file, target_lang)
            
            st.success('处理完成!')
            st.write(f"统计信息:")
            st.write(f"- 总工作表数: {stats['total_sheets']}")
            st.write(f"- 总行数: {stats['total_rows']}")
            st.write(f"- 总单元格数: {stats['total_cells']}")
            st.write(f"- 非目标语言单元格数: {stats['non_target_cells']}")
            st.write(f"- 处理时间: {stats['processing_time']:.2f} 秒")
            
            st.download_button(
                label="下载处理后的Excel文件",
                data=processed_data,
                file_name=f"processed_{uploaded_file.name}",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

if __name__ == "__main__":
    main()