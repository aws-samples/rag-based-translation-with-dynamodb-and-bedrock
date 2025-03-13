from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, BackgroundTasks, status
from fastapi.responses import StreamingResponse
from typing import List, Dict, Any
import io
from io import BytesIO
import openpyxl

from app.models.translation import TranslationRequest, TranslationResponse, FileTranslationRequest, FileTranslationStats
from app.models.user import User
from app.api.dependencies.auth import get_current_active_user
from app.services.translation_service import translate_text, process_excel
from app.services.dictionary_service import list_dictionaries

router = APIRouter()

@router.post("/text", response_model=TranslationResponse)
async def translate_text_endpoint(
    request: TranslationRequest,
    current_user: User = Depends(get_current_active_user)
):
    """
    Translate text content
    """
    try:
        translated_text, term_mapping = await translate_text(
            contents=request.contents,
            source_lang=request.source_lang,
            target_lang=request.target_lang,
            dictionary_id=request.dictionary_id,
            model_id=request.model_id,
            lambda_alias=request.lambda_alias
        )
        
        if translated_text is None:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Translation failed: No translation result returned"
            )
            
        return TranslationResponse(
            translated_text=translated_text,
            term_mapping=term_mapping
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Translation failed: {str(e)}"
        )

@router.post("/file")
async def translate_file_endpoint(
    file: UploadFile = File(...),
    target_lang: str = Form(...),
    dictionary_id: str = Form(...),
    model_id: str = Form(...),
    concurrency: int = Form(3),
    lambda_alias: str = Form("staging"),
    current_user: User = Depends(get_current_active_user)
):
    """
    Translate Excel file content
    """
    # 验证文件名
    if not file.filename:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="文件名不能为空"
        )
        
    # 验证文件扩展名
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="只支持Excel文件格式(.xlsx, .xls)"
        )
    
    try:
        # 读取文件内容
        file_content = await file.read()
        
        # 验证文件内容
        if not file_content or len(file_content) == 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="文件内容为空"
            )
        
        # 验证文件大小
        if len(file_content) > 10 * 1024 * 1024:  # 10MB
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="文件大小不能超过10MB"
            )
            
        # 验证Excel文件格式
        try:
            # 尝试加载工作簿以验证它是否为有效的Excel文件
            file_io = BytesIO(file_content)
            file_io.seek(0)
            try:
                workbook = openpyxl.load_workbook(file_io, read_only=True)
                
                # 验证是否有工作表
                if len(workbook.sheetnames) == 0:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="Excel文件不包含任何工作表"
                    )
                    
                # 检查是否有内容
                has_content = False
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    if sheet.max_row > 0 and sheet.max_column > 0:
                        has_content = True
                        break
                
                if not has_content:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="Excel文件不包含任何内容"
                    )
                    
            except Exception as excel_error:
                if "file is not a zip file" in str(excel_error).lower():
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="无效的Excel文件格式"
                    )
                else:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Excel文件格式错误: {str(excel_error)}"
                    )
        except HTTPException:
            raise
        except Exception as excel_error:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"无法解析Excel文件: {str(excel_error)}"
            )
        
        # 验证模型ID
        from app.services.aws_service import aws_service
        available_models = await aws_service.list_foundation_models()
        if model_id not in available_models:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"不支持的翻译模型: {model_id}"
            )
            
        # 验证目标语言
        from app.core.config import settings
        if target_lang not in settings.SUPPORTED_LANGUAGES.values():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"不支持的目标语言: {target_lang}"
            )
        
        # 重置文件指针
        file_io = BytesIO(file_content)
        file_io.seek(0)
        
        # 处理Excel文件
        processed_data, stats = await process_excel(
            file_content=file_content,
            target_lang=target_lang,
            model_id=model_id,
            dict_id=dictionary_id,
            concurrency=concurrency,
            lambda_alias=lambda_alias
        )
        
        # 处理文件名，确保它是URL安全的
        import urllib.parse
        safe_filename = urllib.parse.quote(file.filename)
        
        # 返回处理后的文件作为下载
        return StreamingResponse(
            io.BytesIO(processed_data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{safe_filename}"
            }
        )
    except HTTPException:
        # 重新抛出HTTP异常
        raise
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"文件翻译错误: {str(e)}\n{error_details}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"文件翻译失败: {str(e)}"
        )

@router.get("/models")
async def list_models(current_user: User = Depends(get_current_active_user)):
    """
    List available translation models from Amazon Bedrock
    """
    from app.services.aws_service import aws_service
    models = await aws_service.list_foundation_models()
    return {"models": models}

@router.get("/languages")
async def list_languages(current_user: User = Depends(get_current_active_user)):
    """
    List supported languages
    """
    from app.core.config import settings
    return {"languages": settings.SUPPORTED_LANGUAGES}

@router.get("/dictionaries")
async def get_dictionaries(current_user: User = Depends(get_current_active_user)):
    """
    List available dictionaries
    """
    try:
        dictionaries = await list_dictionaries()
        return {"dictionaries": dictionaries or ["default_dictionary"]}
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to list dictionaries: {str(e)}"
        )
