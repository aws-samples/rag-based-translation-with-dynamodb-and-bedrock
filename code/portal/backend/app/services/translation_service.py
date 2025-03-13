import asyncio
import time
import openpyxl
from io import BytesIO
from typing import List, Dict, Tuple, Any, Optional

from app.services.aws_service import aws_service

async def translate_text(
    contents: List[str],
    source_lang: str,
    target_lang: str,
    dictionary_id: str,
    model_id: str,
    lambda_alias: str = "prod"
) -> Tuple[str, str]:
    """
    Translate text content
    """
    return await aws_service.translate_content(
        contents=contents,
        source_lang=source_lang,
        target_lang=target_lang,
        dictionary_id=dictionary_id,
        model_id=model_id,
        lambda_alias=lambda_alias
    )

def is_not_number(text: str) -> bool:
    """
    Check if text is not a number
    """
    try:
        float(text)
        return False
    except ValueError:
        return True

async def process_excel(
    file_content: bytes,
    target_lang: str,
    model_id: str,
    dict_id: str,
    concurrency: int = 3,
    lambda_alias: str = "staging"
) -> Tuple[bytes, Dict[str, Any]]:
    """
    Process Excel file and translate cells
    """
    start_time = time.time()
    workbook = openpyxl.load_workbook(BytesIO(file_content))
    total_sheets = len(workbook.sheetnames)
    
    total_rows = 0
    total_cells = 0
    non_target_cells = 0
    processed_cells = 0

    # Count total cells
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_rows = sheet.max_row
        sheet_cells = sheet.max_column * sheet_rows
        total_rows += sheet_rows
        total_cells += sheet_cells

    # Create a semaphore to limit concurrency
    semaphore = asyncio.Semaphore(concurrency)

    async def process_cell(cell):
        nonlocal non_target_cells, processed_cells
        if cell.data_type == 's' and cell.value and is_not_number(cell.value):
            try:
                lang = await aws_service.detect_language(cell.value)
                if lang and lang != target_lang:
                    async with semaphore:
                        try:
                            translation, _ = await aws_service.translate_content(
                                contents=[cell.value],
                                source_lang=lang,
                                target_lang=target_lang,
                                dictionary_id=dict_id,
                                model_id=model_id,
                                lambda_alias=lambda_alias
                            )
                            if translation:
                                cell.value = translation.strip()
                                non_target_cells += 1
                            else:
                                print(f"Translation returned None for text: {cell.value[:50]}...")
                        except Exception as e:
                            print(f"Translation error: {str(e)}")
            except Exception as e:
                print(f"Language detection error: {str(e)}")
        
        processed_cells += 1

    # Process cells
    tasks = []
    for sheet in workbook:
        for row in sheet.iter_rows():
            for cell in row:
                task = asyncio.create_task(process_cell(cell))
                tasks.append(task)
    
    try:
        await asyncio.gather(*tasks)
    except Exception as e:
        print(f"Error during cell processing: {str(e)}")

    # Calculate processing time
    end_time = time.time()
    processing_time = end_time - start_time

    # Save workbook to bytes
    buffer = BytesIO()
    workbook.save(buffer)
    
    # Return processed file and statistics
    return buffer.getvalue(), {
        "total_sheets": total_sheets,
        "total_rows": total_rows,
        "total_cells": total_cells,
        "non_target_cells": non_target_cells,
        "processing_time": processing_time
    }
